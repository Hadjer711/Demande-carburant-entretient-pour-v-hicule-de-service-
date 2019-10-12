
# Create your tests here.
from openpyxl import Workbook, load_workbook
from pdfrw import pdfwriter
#wb = load_workbook("Copie de Demande d'entretien véhicule.xlsx")

# grab the active worksheet
#ws = wb.active

# Data can be assigned directly to cells
#ws['B7'] = 'Hadjer'

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
#ws['B9'] = datetime.datetime.now()

# Save the file
#wb.save("Copie de Demande d'entretien véhicule.xlsx")
import PyPDF2


print("fichier edit")